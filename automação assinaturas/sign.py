import pandas as pd
import pywhatkit as kit
import time
from datetime import datetime
from openpyxl import load_workbook

#configurações
MODO_TESTE = False      # True = NÃO envia | False = envia
PLANILHA = 'assinatura_limpa.xlsx'
ABA_PLANILHA = 'base_dados'
WAIT_TIME = 15
DELAY_BETWEEN = 47
RELATORIO = 'relatorio.csv'

#carregando dados...
def carregar_dados():
    df = pd.read_excel(
        PLANILHA,
        sheet_name=ABA_PLANILHA,
        engine='calamine'
    )

    df['Tipo'] = df['Tipo'].astype(str).str.strip().str.lower()
    df['Status'] = df['Status'].astype(str).str.strip().str.lower()
    df['Telefone'] = (
        df['Telefone']
        .astype(str)
        .str.strip()
        .str.replace(r'\.0$', '', regex=True)
        .str.replace(r'\D', '', regex=True)
    )

    df['Gênero'] = (
        df['Gênero']
        .astype(str)
        .str.strip()
        .str.upper()
    )

    return df
df = carregar_dados()

LINHA_EXCEL = 1150

df = carregar_dados()

df = df.iloc[LINHA_EXCEL - 2:]

#filtro
df_filtrado = df[
    (df['Status'] == '2° envio') &
    (df['Tipo'].str.contains('novo|renovação', na=False)) &
    (df['Telefone'].notna()) &
    (df['Telefone'] != '') &
    (df['Telefone'] != '0')
].copy()

#exibirpython 

log_envios = []

print(df_filtrado[['Nome do cliente', 'Telefone']])

#laço de rep
for i, cliente in df_filtrado.iterrows():

    horario = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        nome = str(cliente['Nome do cliente']).strip()
        telefone = "+" + str(cliente['Telefone']).strip()
        if telefone in [None, '', '0']:
            print(f"[ERRO] telefone inválido: {nome}")
            continue
        seguradora = str(cliente['Seguradora']).strip()
        consultor = str(cliente['Consultor']).strip()
        genero = str(cliente['Gênero']).strip().upper()

        if genero == 'F':
            cargo = "a consultora"
        else:
            cargo = "o consultor"

        mensagem = f"""Olá, *{nome}*! 😊
Sou a Ana, do departamento de assinaturas da corretora *Prontocar Seguros*.
Você fechou seu seguro com {cargo} *{consultor}*, pela *{seguradora}*.

Encaminhei a sua proposta por e-mail. Consegue verificar e assinar, por gentileza?

Fico à disposição para qualquer dúvida!"""

        #exibição envio
        print(f"\n[{horario}] Enviando para {nome} ({telefone})...")

        #envio
        if not MODO_TESTE:
            kit.sendwhatmsg_instantly(
                phone_no=telefone,
                message=mensagem,
                wait_time=WAIT_TIME,
                tab_close=True
            )
            time.sleep(DELAY_BETWEEN)
        else:
            print(">>> MODO TESTE <<<")

        log_envios.append({
            'horario': horario,
            'nome': nome,
            'telefone': telefone,
            'status': 'enviado',
            'erro': ''
        })

        print(f"[{horario}] Processado para {nome}!\n")

    except Exception as e:
        log_envios.append({
            'horario': horario,
            'nome': cliente.get('Nome do cliente', ''),
            'telefone': cliente.get('Telefone', ''),
            'status': 'falhou',
            'erro': str(e)
        })

        print(f"[{horario}] ERRO ao enviar:")
        print(e)

#relatório geral
pd.DataFrame(log_envios).to_csv(
    RELATORIO,
    sep=';',
    index=False, #---> tira coluna que atrapalha leitura csv
    encoding='utf-8-sig' #---> salva caracteres especiais sem erro
)

#atualizar status excel
if not MODO_TESTE:

    enviados_ok = [
        l['telefone']
        for l in log_envios
        if l['status'] == 'enviado'
    ]

    wb = load_workbook(PLANILHA)
    ws = wb[ABA_PLANILHA]

    headers = {}

    for cell in ws[1]:
        headers[cell.value] = cell.column

    col_telefone = headers['Telefone']
    col_status = headers['Status']

    #atualiza somente as células necessárias
    for row in range(2, ws.max_row + 1):

        telefone_excel = str(
            ws.cell(row=row, column=col_telefone).value
        ).split('.')[0]

        if telefone_excel in enviados_ok:

            ws.cell(
                row=row,
                column=col_status
            ).value = 'Mensagem enviada'

    wb.save(PLANILHA)

#resumo
enviadas = len([l for l in log_envios if l['status'] == 'enviado'])
falhas = len([l for l in log_envios if l['status'] == 'falhou'])

print("=== PROCESSO FINALIZADO ===")
print(f"Mensagens processadas: {enviadas}")
print(f"Falhas: {falhas}")

if falhas:
    print("Clientes com erro:")
    for l in log_envios:
        if l['status'] == 'falhou':
            print(f"- {l['nome']} ({l['telefone']}): {l['erro']}")